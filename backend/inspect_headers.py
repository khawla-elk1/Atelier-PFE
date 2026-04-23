import openpyxl
import os

file_path = "Pointage_materiel_cpb_20260101_20260331_20260407134848984_.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    # Print first row (headers)
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        print(f"Headers: {row}")
    # Print first 2 data rows
    for row in sheet.iter_rows(min_row=2, max_row=3, values_only=True):
        print(f"Data: {row}")
except Exception as e:
    print(f"Error: {e}")
