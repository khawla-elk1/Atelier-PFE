import openpyxl
import re

import os

# Chargement du fichier
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = os.path.join(SCRIPT_DIR, 'PARC MATERIEL V 21.02.2026.xlsx')
wb = openpyxl.load_workbook(FILE_NAME, data_only=True)

def esc(val):
    """Echappement pour SQL"""
    if val is None: return ''
    s = str(val).strip()
    s = s.replace("'", "''")
    s = s.replace('\n', ' ').replace('\r', ' ')
    return s[:250]

def date_str(val):
    if val is None: return ''
    s = str(val).strip()
    if '00:00:00' in s: s = s.split(' ')[0]
    if s == '' or s == '-' or 'None' in s or '00:00:00' in s: return ''
    return s[:50]

# Colonnes dans l'onglet 'Global' (d'après l'analyse du fichier)
# Col 0: Matériel (Genre)
# Col 1: Genre Matériel
# Col 2: Marque
# Col 3: Type (Modèle)
# Col 4: Poids
# Col 5: Série chassis
# Col 6: Type moteur
# Col 7: Série moteur
# Col 8: Immatriculation
# Col 9: Date d'acquisition
# Col 10: Date mise en circulation
# Col 11: Code Initiale Matériel
# Col 12: Nouveau code Matériel

sql_lines = [
    "SET FOREIGN_KEY_CHECKS = 0;",
    "-- Import GLOBAL STAPORT BTP 832 items",
    ""
]

seen_codes = set()
count = 0

if 'Global' in wb.sheetnames:
    ws = wb['Global']
    # On commence à la ligne 11 car les lignes 1-10 sont des entêtes
    for row in ws.iter_rows(min_row=11, values_only=True):
        if not any(row): continue
        
        row_list = list(row) + [None]*20 # Padding
        
        genre = esc(row_list[0])
        genre_sub = esc(row_list[1])
        marque = esc(row_list[2])
        modele = esc(row_list[3])
        poids = esc(row_list[4])
        serie_ch = esc(row_list[5])
        type_mot = esc(row_list[6])
        serie_mot = esc(row_list[7])
        immat = esc(row_list[8])
        date_acq = date_str(row_list[9])
        date_mec = date_str(row_list[10])
        code_int = esc(row_list[11])
        code_mat = esc(row_list[12])
        
        # Identification unique (Matricule)
        matricule = code_mat if code_mat else code_int if code_int else immat
        if not matricule or matricule.lower() == 'none' or matricule == '':
            # Si pas de code, on saute la ligne (probablement une ligne vide ou de titre)
            continue
            
        # Catégorisation automatique
        cat = 'Engin'
        g_low = genre.lower() + " " + genre_sub.lower()
        if any(w in g_low for w in ['voiture', 'pick-up', 'dacia', 'toyota', 'bmw', 'audi']):
            cat = 'Voiture'
        elif any(w in g_low for w in ['camion', 'semi', 'benne', 'malaxeur', 'citerne', 'remorque']):
            cat = 'Camion'
            
        # Nettoyage des doublons
        if matricule in seen_codes:
            matricule = f"{matricule}-{count}"
        seen_codes.add(matricule)
        
        unite = 'km' if cat in ['Voiture', 'Camion'] else 'h'
        
        # SQL Injection safe-ish
        sql = f"INSERT INTO engins (matricule, marque, modele, type, categorie, poids, serie_chassis, type_moteur, serie_moteur, immatriculation, date_acquisition, date_mise_en_circulation, code_interne, code_materiel, unite_compteur, compteur_actuel, statut) " \
              f"VALUES ('{esc(matricule)}', '{marque}', '{modele}', '{genre} {genre_sub}', '{cat}', '{poids}', '{serie_ch}', '{type_mot}', '{serie_mot}', '{immat}', '{date_acq}', '{date_mec}', '{code_int}', '{code_mat}', '{unite}', 0, 'ACTIF') " \
              f"ON DUPLICATE KEY UPDATE marque='{marque}', modele='{modele}', type='{genre} {genre_sub}', categorie='{cat}', poids='{poids}', serie_chassis='{serie_ch}', type_moteur='{type_mot}', serie_moteur='{serie_mot}', immatriculation='{immat}', date_acquisition='{date_acq}', date_mise_en_circulation='{date_mec}', code_interne='{code_int}', code_materiel='{code_mat}', statut='ACTIF';"
        sql_lines.append(sql)
        count += 1

OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'main', 'resources', 'data.sql')
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_lines))

print(f"IMPORT REUSSI : {count} lignes extraites de l'onglet Global vers data.sql")
