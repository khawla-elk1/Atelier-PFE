import openpyxl

wb = openpyxl.load_workbook(r'PARC MATERIEL V 21.02.2026.xlsx', data_only=True)
out = open('parc_data.txt', 'w', encoding='utf-8')

out.write(f"SHEETS: {wb.sheetnames}\n\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    out.write(f"{'='*60}\n")
    out.write(f"SHEET: {sheet_name} (rows={ws.max_row}, cols={ws.max_column})\n")
    out.write(f"{'='*60}\n")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(25, ws.max_row), values_only=True), 1):
        cleaned = [str(v) if v is not None else '' for v in row[:20]]
        if any(c.strip() for c in cleaned):
            out.write(f"Row {row_idx}: {' | '.join(cleaned)}\n")
    out.write("\n")

out.close()
print("DONE - saved to parc_data.txt")
